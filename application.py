from flask import Flask, flash, redirect, render_template, request, session
from flask_session import Session
from tempfile import mkdtemp
from werkzeug.exceptions import default_exceptions
import xlsxwriter
from xlrd import open_workbook
from openpyxl import load_workbook
from xlutils.copy import copy
import pandas
import os
import pdb

from helpers import lookupis, lookupbs, lookupcf, incomestatementdict, balancesheetdict, cashflowdict, usd

# Configure application
app = Flask(__name__)

# Ensure responses aren't cached
if app.config["DEBUG"]:
    @app.after_request
    def after_request(response):
        response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
        response.headers["Expires"] = 0
        response.headers["Pragma"] = "no-cache"
        return response

# Custom filter
# app.jinja_env.filters["usd"] = usd

# Configure session to use filesystem (instead of signed cookies)
app.config["SESSION_FILE_DIR"] = mkdtemp()
app.config["SESSION_PERMANENT"] = False
app.config["SESSION_TYPE"] = "filesystem"
Session(app)

@app.route("/", methods=["GET", "POST"])
def form():

    if request.method == "POST":

        # Get user inputs from form
        company = request.form.get("company")
        if company == "":
            return render_template("error.html")

        session['ebitgrowth'] = request.form.get("ebitgrowth")
        if session['ebitgrowth'] == "":
            return render_template("error.html")

        session['depamtgrowth'] = request.form.get("depamtgrowth")
        if session['depamtgrowth'] == "":
            return render_template("error.html")

        session['capexgrowth'] = request.form.get("capexgrowth")
        if session['capexgrowth'] == "":
            return render_template("error.html")

        session['niwcgrowth'] = request.form.get("niwcgrowth")
        if session['niwcgrowth'] == "":
            return render_template("error.html")

        session['discountrate'] = request.form.get("discountrate")

        session['taxrate'] = request.form.get("taxrate")
        if session['taxrate'] == "":
            return render_template("error.html")

        session['shares'] = request.form.get("shares")
        if session['shares'] == "":
            return render_template("error.html")

        # Lookup Income Statement
        incomestatementdict = lookupis(company)

        # Lookup Balance Sheet
        balancesheetdict = lookupbs(company)

        # Lookup Cash Flow
        cashflowdict = lookupcf(company)

        # Return the variables assigned in this function
        return render_template("inputs.html", incomestatementdict=incomestatementdict, balancesheetdict=balancesheetdict, cashflowdict=cashflowdict)

    else:

        return render_template("form.html")

@app.route("/inputs", methods=["GET", "POST"])
def inputs():

    if request.method == "POST":

        # Find ebitindex
        for n in range(len(incomestatementdict)):
            if incomestatementdict[n]['lineitem'] == 'EBITDA':
                ebitindex = n
                break

        # Find depamtindex
        for n in range(len(cashflowdict)):
            if cashflowdict[n]['lineitem'] == 'Depreciation & amortization':
                depamtindex = n
                break

        # Find longtermdebtindex
        for n in range(len(balancesheetdict)):
            if balancesheetdict[n]['lineitem'] == 'Long-term debt':
                longtermdebtindex = n
                break

        # Find shorttermdebtindex
        for n in range(len(balancesheetdict)):
            if balancesheetdict[n]['lineitem'] == 'Short-term debt':
                shorttermdebtindex = n
                break

        # Find cashcashindex
        for n in range(len(balancesheetdict)):
            if balancesheetdict[n]['lineitem'] == 'Cash and cash equivalents':
                cashcashindex = n
                break

        # Find capexindex
        for n in range(len(cashflowdict)):
            if cashflowdict[n]['lineitem'] == 'Capital expenditure':
                capexindex = n
                break

        # Find totalcurrentassetsindex
        for n in range(len(balancesheetdict)):
            if balancesheetdict[n]['lineitem'] == 'Total current assets':
                totalcurrentassetsindex = n
                break

        # Find totalcurrentliabilitiesindex
        for n in range(len(balancesheetdict)):
            if balancesheetdict[n]['lineitem'] == 'Total current liabilities':
                totalcurrentliabilitiesindex = n
                break


        # Pull relevant inputs from income statement, balance sheet, cash flow statement
        ebit = int(incomestatementdict[ebitindex]["yr5"])-int(cashflowdict[depamtindex]["yr5"])
        if balancesheetdict[longtermdebtindex]["yr5"] == "":
            balancesheetdict[longtermdebtindex]["yr5"] = "0"
        else:
            balancesheetdict[longtermdebtindex]["yr5"] = balancesheetdict[longtermdebtindex]["yr5"]
        if balancesheetdict[shorttermdebtindex]["yr5"] == "":
            balancesheetdict[shorttermdebtindex]["yr5"] = "0"
        else:
            balancesheetdict[longtermdebtindex]["yr5"] = balancesheetdict[longtermdebtindex]["yr5"]
        netdebt = (int(balancesheetdict[longtermdebtindex]["yr5"])+int(balancesheetdict[shorttermdebtindex]["yr5"]))-int(balancesheetdict[cashcashindex]["yr5"])
        depamt = int(cashflowdict[depamtindex]["yr5"])
        #Capital Expenditure
        capex = int(cashflowdict[capexindex]["yr5"])
        # Net Increase in Working Capital
        niwc = (int(balancesheetdict[totalcurrentassetsindex]["yr5"])-int(balancesheetdict[totalcurrentliabilitiesindex]["yr5"]))-(int(balancesheetdict[totalcurrentassetsindex]["yr4"])-int(balancesheetdict[totalcurrentliabilitiesindex]["yr4"]))
        # Predicted EBIT Growth
        ebitgrowth = float(session['ebitgrowth'])
        # Predicted Dep & Amt Growth
        depamtgrowth = float(session['depamtgrowth'])
        # Predicted CapEx growth
        capexgrowth = float(session['capexgrowth'])
        # Predicted NIWC growth
        niwcgrowth = float(session['niwcgrowth'])
        # Discount Rate
        discountrate = float(session['discountrate'])
        # Corporate Tax Rate
        taxrate = float(session['taxrate'])
        # Number of Shares Outstanding
        shares = float(session['shares'])

        # Build DCF Model using xlsxwriter

        # Create a new Excel Workbook
        workbook = xlsxwriter.Workbook('dcf.xlsx')

        # Add a Worksheet
        worksheetdcf = workbook.add_worksheet('dcf')

        #Build DCF Model
        worksheetdcf.write('A1', 'Unlevered Free Cash Flow Calculation')
        worksheetdcf.write('A2', 'All numbers in Millions except share data')
        worksheetdcf.write('A6', 'EBIT')
        worksheetdcf.write('E6', 'EBIT Growth')
        worksheetdcf.write('F6', ebitgrowth)
        worksheetdcf.write('A7', 'Plus: Non-deductible Goodwill Amort.')
        worksheetdcf.write('A9', 'Less: Provision for Taxes')
        worksheetdcf.write('A12', 'Plus: D&A (excl. non-deductible GW amort.)')
        worksheetdcf.write('A13', 'Less: Capital Expenditures')
        worksheetdcf.write('A14', 'Less: Increase in Net Working Capital')
        worksheetdcf.write('E12', 'D&A Growth')
        worksheetdcf.write('F12', depamtgrowth)
        worksheetdcf.write('E13', 'CapEx Growth')
        worksheetdcf.write('F13', capexgrowth)
        worksheetdcf.write('E14', 'NIWC Growth')
        worksheetdcf.write('F14', niwcgrowth)
        worksheetdcf.write('B8', 'EBITDA')
        worksheetdcf.write('B10', 'Unlevered Net Income')
        worksheetdcf.write('B15', 'Unlevered Free Cash Flow')
        worksheetdcf.write('H3', 'Calendar Year Ending December 31,')
        worksheetdcf.write('H4', '2016A')
        worksheetdcf.write('J4', '2017P')
        worksheetdcf.write('L4', '2018P')
        worksheetdcf.write('N4', '2019E')
        worksheetdcf.write('P4', '2020E')
        worksheetdcf.write('R4', '2021E')
        worksheetdcf.write('H6', ebit)
        worksheetdcf.write('H12', depamt)
        worksheetdcf.write('H13', capex)
        worksheetdcf.write('H14', niwc)

        worksheetdcf.write('J6', '=H6*(1+F6)')
        worksheetdcf.write('L6', '=J6*(1+F6)')
        worksheetdcf.write('N6', '=L6*(1+F6)')
        worksheetdcf.write('P6', '=N6*(1+F6)')
        worksheetdcf.write('R6', '=P6*(1+F6)')

        worksheetdcf.write('J12', '=H12*(1+F12)')
        worksheetdcf.write('L12', '=J12*(1+F12)')
        worksheetdcf.write('N12', '=L12*(1+F12)')
        worksheetdcf.write('P12', '=N12*(1+F12)')
        worksheetdcf.write('R12', '=P12*(1+F12)')

        worksheetdcf.write('J13', '=H13*(1+F13)')
        worksheetdcf.write('L13', '=J13*(1+F13)')
        worksheetdcf.write('N13', '=L13*(1+F13)')
        worksheetdcf.write('P13', '=N13*(1+F13)')
        worksheetdcf.write('R13', '=P13*(1+F13)')

        worksheetdcf.write('J14', '=H14*(1+F14)')
        worksheetdcf.write('L14', '=J14*(1+F14)')
        worksheetdcf.write('N14', '=L14*(1+F14)')
        worksheetdcf.write('P14', '=N14*(1+F14)')
        worksheetdcf.write('R14', '=P14*(1+F14)')

        worksheetdcf.write('H8', '=H6+H7')
        worksheetdcf.write('J8', '=J6+J7')
        worksheetdcf.write('L8', '=L6+L7')
        worksheetdcf.write('N8', '=N6+N7')
        worksheetdcf.write('P8', '=P6+P7')
        worksheetdcf.write('R8', '=R6+R7')

        worksheetdcf.write('H9', '=-H8*$X$16')
        worksheetdcf.write('J9', '=-J8*$X$16')
        worksheetdcf.write('L9', '=-L8*$X$16')
        worksheetdcf.write('N9', '=-N8*$X$16')
        worksheetdcf.write('P9', '=-P8*$X$16')
        worksheetdcf.write('R9', '=-R8*$X$16')

        worksheetdcf.write('H10', '=H8+H9')
        worksheetdcf.write('J10', '=SUM(J8:J9)')
        worksheetdcf.write('L10', '=SUM(L8:L9)')
        worksheetdcf.write('N10', '=SUM(N8:N9)')
        worksheetdcf.write('P10', '=SUM(P8:P9)')
        worksheetdcf.write('R10', '=SUM(R8:R9)')

        worksheetdcf.write('H15', '=SUM(H10:H14)')
        worksheetdcf.write('J15', '=SUM(J10:J14)')
        worksheetdcf.write('L15', '=SUM(L10:L14)')
        worksheetdcf.write('N15', '=SUM(N10:N14)')
        worksheetdcf.write('P15', '=SUM(P10:P14)')
        worksheetdcf.write('R15', '=SUM(R10:R14)')

        worksheetdcf.write('A19', 'DCF Analysis (2017-2021): Perpetuity Growth Method')

        worksheetdcf.write('F21', 'Total Enterprise Value')
        worksheetdcf.write('F22', 'Terminal Perpetuity Growth Rate')
        worksheetdcf.write('F23', ebitgrowth - 0.2*(ebitgrowth))
        worksheetdcf.write('H23', ebitgrowth)
        worksheetdcf.write('J23', ebitgrowth + 0.2*(ebitgrowth))
        worksheetdcf.write('C24', 'Discount')
        worksheetdcf.write('C25', 'Rate')
        worksheetdcf.write('C26', '(WACC)')
        worksheetdcf.write('D24', discountrate)
        worksheetdcf.write('D25', discountrate + 0.01)
        worksheetdcf.write('D26', discountrate + 0.02)
        worksheetdcf.write('F24', '=NPV($D24,$J$15:$P$15,$R$15+$R$15*(1+F$23)/($D24-F$23))')
        worksheetdcf.write('H24', '=NPV($D24,$J$15:$P$15,$R$15+$R$15*(1+H$23)/($D24-H$23))')
        worksheetdcf.write('J24', '=NPV($D24,$J$15:$P$15,$R$15+$R$15*(1+J$23)/($D24-J$23))')
        worksheetdcf.write('F25', '=NPV($D25,$J$15:$P$15,$R$15+$R$15*(1+F$23)/($D25-F$23))')
        worksheetdcf.write('H25', '=NPV($D25,$J$15:$P$15,$R$15+$R$15*(1+H$23)/($D25-H$23))')
        worksheetdcf.write('J25', '=NPV($D25,$J$15:$P$15,$R$15+$R$15*(1+J$23)/($D25-J$23))')
        worksheetdcf.write('F26', '=NPV($D26,$J$15:$P$15,$R$15+$R$15*(1+F$23)/($D26-F$23))')
        worksheetdcf.write('H26', '=NPV($D26,$J$15:$P$15,$R$15+$R$15*(1+H$23)/($D26-H$23))')
        worksheetdcf.write('J26', '=NPV($D26,$J$15:$P$15,$R$15+$R$15*(1+J$23)/($D26-J$23))')

        worksheetdcf.write('P21', 'Total Equity Value')
        worksheetdcf.write('P22', 'Terminal Perpetuity Growth Rate')
        worksheetdcf.write('P23', '=F23')
        worksheetdcf.write('R23', '=H23')
        worksheetdcf.write('T23', '=J23')
        worksheetdcf.write('M24', 'Discount')
        worksheetdcf.write('M25', 'Rate')
        worksheetdcf.write('M26', '(WACC)')
        worksheetdcf.write('N24', '=D24')
        worksheetdcf.write('N25', '=D25')
        worksheetdcf.write('N26', '=D26')
        worksheetdcf.write('P24', '=F24-$X$17')
        worksheetdcf.write('R24', '=H24-$X$17')
        worksheetdcf.write('T24', '=J24-$X$17')
        worksheetdcf.write('P25', '=F25-$X$17')
        worksheetdcf.write('R25', '=H25-$X$17')
        worksheetdcf.write('T25', '=J25-$X$17')
        worksheetdcf.write('P26', '=F26-$X$17')
        worksheetdcf.write('R26', '=H26-$X$17')
        worksheetdcf.write('T26', '=J26-$X$17')

        worksheetdcf.write('F28', 'Implied Terminal EBITDA Multiple')
        worksheetdcf.write('F29', 'Terminal Perpetuity Growth Rate')
        worksheetdcf.write('F30', '=F23')
        worksheetdcf.write('H30', '=H23')
        worksheetdcf.write('J30', '=J23')
        worksheetdcf.write('C31', 'Discount')
        worksheetdcf.write('C32', 'Rate')
        worksheetdcf.write('C33', '(WACC)')
        worksheetdcf.write('D31', '=D24')
        worksheetdcf.write('D32', '=D25')
        worksheetdcf.write('D33', '=D26')
        worksheetdcf.write('F31', '=$R$15*(1+F$30)/($D31-F$30)/$R$6')
        worksheetdcf.write('H31', '=$R$15*(1+H$30)/($D31-H$30)/$R$6')
        worksheetdcf.write('J31', '=$R$15*(1+J$30)/($D31-J$30)/$R$6')
        worksheetdcf.write('F32', '=$R$15*(1+F$30)/($D32-F$30)/$R$6')
        worksheetdcf.write('H32', '=$R$15*(1+H$30)/($D32-H$30)/$R$6')
        worksheetdcf.write('J32', '=$R$15*(1+J$30)/($D32-J$30)/$R$6')
        worksheetdcf.write('F33', '=$R$15*(1+F$30)/($D33-F$30)/$R$6')
        worksheetdcf.write('H33', '=$R$15*(1+H$30)/($D33-H$30)/$R$6')
        worksheetdcf.write('J33', '=$R$15*(1+J$30)/($D33-J$30)/$R$6')

        worksheetdcf.write('P28', 'Total Price Per Share')
        worksheetdcf.write('P29', 'Terminal Perpetuity Growth Rate')
        worksheetdcf.write('P30', '=F23')
        worksheetdcf.write('R30', '=H23')
        worksheetdcf.write('T30', '=J23')
        worksheetdcf.write('M31', 'Discount')
        worksheetdcf.write('M32', 'Rate')
        worksheetdcf.write('M33', '(WACC)')
        worksheetdcf.write('N31', '=D24')
        worksheetdcf.write('N32', '=D25')
        worksheetdcf.write('N33', '=D26')
        worksheetdcf.write('P31', '=P24/$X$18')
        worksheetdcf.write('R31', '=R24/$X$18')
        worksheetdcf.write('T31', '=T24/$X$18')
        worksheetdcf.write('P32', '=P25/$X$18')
        worksheetdcf.write('R32', '=R25/$X$18')
        worksheetdcf.write('T32', '=T25/$X$18')
        worksheetdcf.write('P33', '=P26/$X$18')
        worksheetdcf.write('R33', '=R26/$X$18')
        worksheetdcf.write('T33', '=T26/$X$18')

        worksheetdcf.write('V16', 'Tax Rate')
        worksheetdcf.write('V17', 'Net Debt')
        worksheetdcf.write('V18', 'Shares')
        worksheetdcf.write('X16', taxrate)
        worksheetdcf.write('X17', netdebt)
        worksheetdcf.write('X18', shares)

        workbook.close()

        # Use openpyxlsx library to read stock price value
        wb = load_workbook(filename='dcf.xlsx', read_only=True)
        ws = wb['dcf']

        bear = ws['P31'].value
        base = ws['R31'].value
        bull = ws['T31'].value

        return render_template("result.html", bear=bear, base=base, bull=bull)
    else:
        return render_template("error.html")
